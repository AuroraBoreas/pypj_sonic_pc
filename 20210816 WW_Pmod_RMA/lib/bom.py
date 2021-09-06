"""
~ transfer the wait-to-maintain pmods to "联络书（新品维护）ver2 210730.xlsx"
  >>> write the wait-to-maintain pmod list to it

author
~ @ZL, 202010803

changelog
~ v0.01, initial build

"""

import pandas as pd
from openpyxl import load_workbook
from lib.quotation import Path, DataFrame, logging

def remove_history(ws):
    for row in ws.iter_rows(min_row=19, min_col=2, max_row=38, max_col=7):
        for cell in row:
            cell.value = None

def generate_bom_excel(template:Path, df:DataFrame, dst:Path)->None:
    sh_name = '210201'
    dst_start_row = 18 # python index starts from 0
    dst_start_col = 1  # python index starts from 0

    wb = load_workbook(template)
    with pd.ExcelWriter(template, engine='openpyxl', mode='r') as writer:
        writer.book   = wb
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        try:
            remove_history(wb[sh_name])
            df.to_excel(writer, sheet_name=sh_name, startrow=dst_start_row, startcol=dst_start_col, index=False, header=False)
            wb.save(dst)
            remove_history(wb[sh_name])
            logging.info('New BOM(Excel) created successed!')
        except AttributeError:
            logging.info('Failed: the new bom registration list length exceeded 10 rows')
        except PermissionError:
            logging.info('Failed: the template is opened')