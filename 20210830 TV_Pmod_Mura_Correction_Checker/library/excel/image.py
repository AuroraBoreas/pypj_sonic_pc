"""
this module loads Mura Images into an excel template;

it has the following functionalities
- sort images from a given source image folder
- open template excel workbook
- load resized images into specific columns in the workbook

author
- @ZL, 20210828

changelog
- v0.01, initial build

Ref links:
- https://stackoverflow.com/questions/51591634/fill-a-image-in-a-cell-excel-using-python
- https://stackoverflow.com/questions/2232742/does-python-pil-resize-maintain-the-aspect-ratio
- https://stackoverflow.com/questions/1405602/how-to-adjust-the-quality-of-a-resized-image-in-python-imaging-library

"""

import openpyxl, os, contextlib
from PIL import Image
from typing import Callable, NewType, List
Path = NewType('Path', str)
import logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s %(message)s')
import pathlib, pprint

@contextlib.contextmanager
def open_workbook(wb_path:Path, dst_path:Path):
    wb = openpyxl.load_workbook(wb_path)
    try:
        yield wb
    finally:
        wb.save(dst_path)
        wb.close()

class MuraImageLoader:
    img_ext       = '.bmp'
    idn_muradata  = ('MesData896_', 'MesData2_')
    idn_muradataR = ('MesData896Result_', 'MesData2Result_')
    dstStartRow   = 2
    img_quality   = 95

    def __init__(self, xl_template:Path, dst_xl:Path, img_folder:Path, width:int, dstCellName:str, dstCellColO:str, dstCellColR:str):
        self._xl_template = xl_template
        self._dst_xl      = dst_xl
        self._img_folder  = img_folder
        self.width        = width
        self.dstCellName  = dstCellName
        self.dstCellColO  = dstCellColO
        self.dstCellColR  = dstCellColR
        self._Original:List[Path]  = []
        self._Result:List[Path]    = []

    def _sort_images(self)->None:
        for root, _, files in os.walk(self._img_folder):
            for file in files:
                if file.endswith(self.img_ext):
                    filepath = os.path.join(root, file)
                    if file.startswith(self.idn_muradata):
                        self._Original.append(filepath)
                    if file.startswith(self.idn_muradataR):
                        self._Result.append(filepath)
        logging.info("Before/After QTY::SRC_Images, MesDataO = {}, MesDataR = {}".format(len(self._Original), len(self._Result)))

    def _is_pmod_id_match(self, x:str, y:str)->bool:
        delimiter = '_'
        idx_id    = 2
        return x.split(delimiter)[idx_id] == y.split(delimiter)[idx_id]

    def cmp_before(self)->None:
        idx_w, idx_h = (0, 1)
        i = self.dstStartRow
        k = 0

        with open_workbook(self._xl_template, self._dst_xl) as wb:
            ws = wb.worksheets[0]
            ub = len(self._Original)
            for k in range(0, ub):
                img_896  = self._Original[k]

                img = Image.open(img_896)
                width_percent = (self.width/float(img.size[idx_w]))
                hsize = int((float(img.size[idx_h])*float(width_percent)))
                img = img.resize((self.width, hsize), Image.ANTIALIAS)
                pmod_name = os.path.split(img_896)[-1]
                img_addr = f'tmp_{pmod_name}.png'
                img.save(img_addr, quality=self.img_quality)
                img896 = openpyxl.drawing.image.Image(img_addr)
                dstCellAddress = f'{self.dstCellName}{i}'
                ws[dstCellAddress].value = pmod_name # pmod/SET name
                dstCellAddress = f'{self.dstCellColO}{i}'
                ws.add_image(img896, dstCellAddress)
                i += 1

    def cmp_after(self)->None:
        idx_w, idx_h = (0, 1)
        j = self.dstStartRow
        k = 0

        with open_workbook(self._xl_template, self._dst_xl) as wb:
            ws = wb.worksheets[0]
            ub = len(self._Result)
            for k in range(0, ub):
                img_896R  = self._Result[k]
                
                img = Image.open(img_896R)
                width_percent = (self.width/float(img.size[idx_w]))
                hsize = int((float(img.size[idx_h])*float(width_percent)))
                img = img.resize((self.width, hsize), Image.ANTIALIAS)
                pmod_name = os.path.split(img_896R)[-1]
                img_addr = f'tmp_{pmod_name}.png'
                img.save(img_addr, quality=self.img_quality)
                img896 = openpyxl.drawing.image.Image(img_addr)
                dstCellAddress = f'{self.dstCellName}{j}'
                ws[dstCellAddress].value = pmod_name # pmod/SET name
                dstCellAddress = f'{self.dstCellColR}{j}'
                ws.add_image(img896, dstCellAddress)
                j += 1

    def cmp_both(self)->None:
        idx_w, idx_h = (0, 1)
        i = j = self.dstStartRow
        k = 0
        with open_workbook(self._xl_template, self._dst_xl) as wb:
            ws = wb.worksheets[0]
            ub = len(self._Original) or len(self._Result)
            try:
                for k in range(0, ub):
                    img_896  = self._Original[k]
                    img_896R = self._Result[k]
                    if not self._is_pmod_id_match(img_896, img_896R):
                        raise IndexError(f"Pair {self.muradata896}:{self.muradataresult896}, PmodID Not Match")
                    
                    img = Image.open(img_896)
                    width_percent = (self.width/float(img.size[idx_w]))
                    hsize = int((float(img.size[idx_h])*float(width_percent)))
                    img = img.resize((self.width, hsize), Image.ANTIALIAS)
                    pmod_name = os.path.split(img_896)[-1]
                    img_addr = f'tmp_{pmod_name}.png'
                    img.save(img_addr, quality=self.img_quality)
                    img896 = openpyxl.drawing.image.Image(img_addr)
                    dstCellAddress = f'{self.dstCellName}{i}'
                    ws[dstCellAddress].value = pmod_name # pmod/SET name
                    dstCellAddress = f'{self.dstCellColO}{i}'
                    ws.add_image(img896, dstCellAddress)
                    i += 1

                    img = Image.open(img_896R)
                    img = img.resize((self.width, hsize), Image.ANTIALIAS)
                    pmod_name = os.path.split(img_896R)[-1]
                    img_addr = f'tmp_R{pmod_name}.png'
                    img.save(img_addr, quality=self.img_quality)
                    img896R = openpyxl.drawing.image.Image(img_addr)
                    dstCellAddress = f'{self.dstCellColR}{j}'
                    ws.add_image(img896R, dstCellAddress)
                    j += 1
            except IndexError:
                ub = len(self._Result)
                for m in range(k, ub):
                    img_896R  = self._Result[m]
                    
                    img = Image.open(img_896R)
                    width_percent = (self.width/float(img.size[idx_w]))
                    hsize = int((float(img.size[idx_h])*float(width_percent)))
                    img = img.resize((self.width, hsize), Image.ANTIALIAS)
                    pmod_name = os.path.split(img_896R)[-1]
                    img_addr = f'tmp_{pmod_name}.png'
                    img.save(img_addr, quality=self.img_quality)
                    img896 = openpyxl.drawing.image.Image(img_addr)
                    dstCellAddress = f'{self.dstCellName}{j}'
                    ws[dstCellAddress].value = pmod_name # pmod/SET name
                    dstCellAddress = f'{self.dstCellColR}{j}'
                    ws.add_image(img896, dstCellAddress)
                    j += 1

    def _export(self, f:Callable)->None:
        if f is None:
            raise NotImplementedError()
        f()

    def clean_tmp_imgs(self):
        p = pathlib.Path('.')
        for f in p.glob('*.bin.bmp.png'):
            pathlib.Path.unlink(f)

    def clean(self)->None:
        self._Original.clear()
        self._Result.clear()
        self.clean_tmp_imgs()

    def work(self, f:Callable)->None:
        self._sort_images()
        self._export(f)
        self.clean()

if __name__ == '__main__':
    template       = r'D:\pj_00_codelib\2019_pypj\20210830 TV_Pmod_Mura_Correction_Checker\templates\Mura_List_Template.xlsx'
    dst_xl         = r'D:\pj_00_codelib\2019_pypj\20210830 TV_Pmod_Mura_Correction_Checker\test1.xlsx'
    src_img_folder = r'D:\pj_00_codelib\2019_pypj\20210830 TV_Pmod_Mura_Correction_Checker\Images'
    dstWidth       = 240
    dstCellName    = 'B'
    dstCellColO  = 'C'
    dstCellColR = 'D'
    mil = MuraImageLoader(template, dst_xl, src_img_folder, dstWidth, dstCellName, dstCellColO, dstCellColR)
    mil.work(mil.cmp_both)
    mil.clean()